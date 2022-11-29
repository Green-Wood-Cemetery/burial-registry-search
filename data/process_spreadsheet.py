import json
import re
import dateparser
import requests
import logging
from nameparser import HumanName
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, PatternFill
from openpyxl.comments import Comment
import argparse
import sqlite3
from inflection import parameterize
from os import path
import time
import googlemaps
import sys
sys.path.insert(0, '..')
from lib.interment import Interment

# command line arguments
parser = argparse.ArgumentParser(description='Spreadsheet to JSON')
parser.add_argument('-input', type=str, help='spreadsheet in xslx format')
parser.add_argument('-key', type=str, help='google api key')
parser.add_argument('-sheet', type=str, help='worksheet to transform')
parser.add_argument('-vol', type=int, help='registry volume number')
parser.add_argument('-row_start', type=int, help='row to start parsing, skipping header rows')
parser.add_argument('-marital_status_cols', type=int, help='number of columns used to indicate marital status')
parser.add_argument('-lookup_years', type=str, help='set to "Y" if no interment or death year columns are available')
parser.add_argument('-year_prefix', type=str, help='first two digits of year to prepend to any year abbreviations')
args = parser.parse_args()
workbook = load_workbook(filename=args.input, data_only=True)

# sheet = workbook.active
sheet = workbook[args.sheet]
'''
=====================================================================================================================
SPREADSHEET COLUMN CONSTANTS

Default Configuration
---------------------
A -> IMAGE_FILENAME
B -> INTERMENT_ID
C -> INTERMENT_MONTH
D -> INTERMENT_DAY
E -> INTERMENT_YEAR
F -> NAME
G -> LOT_LOCATION
H -> GRAVE_LOCATION
I -> BIRTH_PLACE
J -> AGE_YEARS
K -> AGE_MONTHS
L -> AGE_DAYS
M -> MARITAL_STATUS_MARRIED
N -> RESIDENCE_CITY
O -> RESIDENCE_STREET
P -> DEATH_LOCATION
Q -> DEATH_MONTH
R -> DEATH_DAY
S -> DEATH_YEAR
T -> CAUSE_OF_DEATH
U -> UNDERTAKER
V -> NOTES
W -> HAS_DIAGRAM
X -> NEEDS_REVIEW 
=====================================================================================================================
'''

# Volumes: 4
#   - Duplicate 'Street & Number' column
#   - One column for marital status
if args.vol == 4:
    IMAGE_FILENAME = 0
    INTERMENT_ID = 1
    INTERMENT_MONTH = 2
    INTERMENT_DAY = 3
    INTERMENT_YEAR = 4
    NAME = 5
    LOT_LOCATION = 6
    GRAVE_LOCATION = 7
    BIRTH_PLACE = 8
    AGE_YEARS = 9
    AGE_MONTHS = 10
    AGE_DAYS = 11
    MARITAL_STATUS_MARRIED = 12
    MARITAL_STATUS_SINGLE = 12
    MARITAL_STATUS_SINGLE_COLUMN = 12
    RESIDENCE_CITY = 13
    RESIDENCE_STREET = 14
    DEATH_LOCATION = 16
    DEATH_DAY = 17
    DEATH_MONTH = 18
    DEATH_YEAR = 19
    CAUSE_OF_DEATH = 20
    UNDERTAKER = 21
    NOTES = 22
    HAS_DIAGRAM = 23
    NEEDS_REVIEW = 24

# Volumes: 16-20
#   - No INTERMENT_YEAR and DEATH_YEAR columns
#   - Two marital status columns (married and single)
elif 16 <= args.vol <= 20:
    IMAGE_FILENAME = 0
    INTERMENT_ID = 1
    INTERMENT_MONTH = 2
    INTERMENT_DAY = 3
    NAME = 4
    LOT_LOCATION = 5
    GRAVE_LOCATION = 6
    BIRTH_PLACE = 7
    AGE_YEARS = 8
    AGE_MONTHS = 9
    AGE_DAYS = 10
    MARITAL_STATUS_MARRIED = 11
    MARITAL_STATUS_SINGLE = 12
    RESIDENCE_CITY = 13
    RESIDENCE_STREET = 14
    DEATH_LOCATION = 15
    DEATH_MONTH = 16
    DEATH_DAY = 17
    CAUSE_OF_DEATH = 18
    UNDERTAKER = 19
    NOTES = 20
    HAS_DIAGRAM = 21
    NEEDS_REVIEW = 22

# Volumes: 5-15, 21-28
#  - Two marital status columns (married and single)
elif 5 <= args.vol <= 15 or 21 <= args.vol <= 28:
    IMAGE_FILENAME = 0
    INTERMENT_ID = 1
    INTERMENT_MONTH = 2
    INTERMENT_DAY = 3
    INTERMENT_YEAR = 4
    NAME = 5
    LOT_LOCATION = 6
    GRAVE_LOCATION = 7
    BIRTH_PLACE = 8
    AGE_YEARS = 9
    AGE_MONTHS = 10
    AGE_DAYS = 11
    MARITAL_STATUS_MARRIED = 12
    MARITAL_STATUS_SINGLE = 13
    RESIDENCE_CITY = 14
    RESIDENCE_STREET = 15
    DEATH_LOCATION = 16
    DEATH_MONTH = 17
    DEATH_DAY = 18
    DEATH_YEAR = 19
    CAUSE_OF_DEATH = 20
    UNDERTAKER = 21
    NOTES = 22
    HAS_DIAGRAM = 23
    NEEDS_REVIEW = 24

# (Default Configuration)
# Volumes: 1-2, 29-60
# - One marital status column
else:
    IMAGE_FILENAME = 0
    INTERMENT_ID = 1
    INTERMENT_MONTH = 2
    INTERMENT_DAY = 3
    INTERMENT_YEAR = 4
    NAME = 5
    LOT_LOCATION = 6
    GRAVE_LOCATION = 7
    BIRTH_PLACE = 8
    AGE_YEARS = 9
    AGE_MONTHS = 10
    AGE_DAYS = 11
    MARITAL_STATUS_MARRIED = 12
    MARITAL_STATUS_SINGLE = 12
    MARITAL_STATUS_SINGLE_COLUMN = 12
    RESIDENCE_CITY = 13
    RESIDENCE_STREET = 14
    DEATH_LOCATION = 15
    DEATH_MONTH = 16
    DEATH_DAY = 17
    DEATH_YEAR = 18
    CAUSE_OF_DEATH = 19
    UNDERTAKER = 20
    NOTES = 21
    HAS_DIAGRAM = 22
    NEEDS_REVIEW = 23

# =====================================================================================================================
# MAIN
# =====================================================================================================================
interments = []

year_prefix = "19"
if args.year_prefix:
    year_prefix = args.year_prefix

count = 2
max = 999999
# Using the values_only because you want to return the cells' values

_see_r_column_regex = re.compile(r'see r column', re.IGNORECASE)

previous = Interment()

wb = Workbook()
ws = wb.active

# EXCEL HEADERS
# create the style template
header = NamedStyle(name="header")
header.font = Font(bold=True)
header.alignment = Alignment(horizontal="center")
header.border = Border(bottom=Side(border_style="thin"))
header.fill = PatternFill("solid", fgColor="D3D3D3")

content = NamedStyle(name="content")
content.alignment = Alignment(horizontal="left")

ws.append([
    "Interment ID\n[TRANSCRIBED]",
    "Image Number\n[TRANSCRIBED]",
    "Interment Date: Month\n[TRANSCRIBED]",
    "Interment Date: Day\n[TRANSCRIBED]",
    "Interment Date: Year\n[TRANSCRIBED]",
    "Interment Date\n[DISPLAYED]",
    "Interment Date - ISO\n [SEARCH]",
    "Name\n[TRANSCRIBED]",
    "Name\n[DISPLAYED]",
    "Name - Last\n[SEARCH]",
    "Name - First\n[SEARCH]",
    "Name - Middle\n[SEARCH]",
    "Name - Salutation\n[SEARCH]",
    "Name - Suffix\n[SEARCH]",
    "Name: Lot Owner\n[DISPLAYED]",
    "Name: Gender Guess\n[SEARCH]",
    "Lot\n[TRANSCRIBED]",
    "Lot: Current\n[DISPLAYED]",
    "Lot: Previous\n[DISPLAYED]",
    "Grave\n[TRANSCRIBED]",
    "Grave: Current\n[DISPLAYED]",
    "Grave: Previous\n[DISPLAYED]",
    "Birth Place\n[TRANSCRIBED]",
    "Birth Place\n[DISPLAYED]",
    "Birth Place: Geocoded\n[GEOCODE]",
    "Birth Place: Faulty Geocode\n[ADMIN]",
    "Birth Place: Street Number\n[GEOCODE]",
    "Birth Place: Street Name Long\n[GEOCODE]",
    "Birth Place: Street Name Short\n[GEOCODE]",
    "Birth Place: Neighborhood\n[GEOCODE]",
    "Birth Place: City\n[GEOCODE]",
    "Birth Place: County\n[GEOCODE]",
    "Birth Place: State Short\n[GEOCODE]",
    "Birth Place: State Long\n[GEOCODE]",
    "Birth Place: Country Long\n[GEOCODE]",
    "Birth Place: Country Short\n[GEOCODE]",
    "Birth Place: Zip Code\n[GEOCODE]",
    "Birth Place: Place ID\n[GEOCODE]",
    "Birth Place: Formatted Address\n[GEOCODE]",
    "Birth Place: Latitude & Longitude\n[GEOCODE]",
    "Age: Years\n[TRANSCRIBED]",
    "Age: Months\n[TRANSCRIBED]",
    "Age: Days\n[TRANSCRIBED]",
    "Age: Hours\n[TRANSCRIBED]",
    "Age: Display\n[DISPLAYED]",
    "Age: Years\n[SEARCH]",
    "Age: Months\n[SEARCH]",
    "Age: Days\n[SEARCH]",
    "Age: Hours\n[SEARCH]",
    "Marital Status: Married\n[TRANSCRIBED]",
    "Marital Status: Single\n[TRANSCRIBED]",
    "Marital Status: Merged\n[DISPLAYED]",
    "Late Residence: City\n[TRANSCRIBED]",
    "Late Residence: City\n[DISPLAYED]",
    "Late Residence: Street\n[TRANSCRIBED]",
    "Late Residence: Street\n[DISPLAYED]",
    "Late Residence: Geocoded\n[GEOCODE]",
    "Late Residence: Faulty Geocode\n[ADMIN]",
    "Late Residence: Street Number\n[GEOCODE]",
    "Late Residence: Street Name Long\n[GEOCODE]",
    "Late Residence: Street Name Short\n[GEOCODE]",
    "Late Residence: Neighborhood\n[GEOCODE]",
    "Late Residence: City\n[GEOCODE]",
    "Late Residence: County\n[GEOCODE]",
    "Late Residence: State Short\n[GEOCODE]",
    "Late Residence: State Long\n[GEOCODE]",
    "Late Residence: Country Long\n[GEOCODE]",
    "Late Residence: Country Short\n[GEOCODE]",
    "Late Residence: Zip Code\n[GEOCODE]",
    "Late Residence: Place ID\n[GEOCODE]",
    "Late Residence: Formatted Address\n[GEOCODE]",
    "Late Residence: Latitude & Longitude\n[GEOCODE]",
    "Place of Death\n[TRANSCRIBED]",
    "Place of Death\n[DISPLAYED]",
    "Place of Death: Geocoded\n[GEOCODE]",
    "Place of Death: Faulty Geocode\n[ADMIN]",
    "Place of Death: Street Number\n[GEOCODE]",
    "Place of Death: Street Name Long\n[GEOCODE]",
    "Place of Death: Street Name Short\n[GEOCODE]",
    "Place of Death: Neighborhood\n[GEOCODE]",
    "Place of Death: City\n[GEOCODE]",
    "Place of Death: County\n[GEOCODE]",
    "Place of Death: State Short\n[GEOCODE]",
    "Place of Death: State Long\n[GEOCODE]",
    "Place of Death: Country Long\n[GEOCODE]",
    "Place of Death: Country Short\n[GEOCODE]",
    "Place of Death: Zip Code\n[GEOCODE]",
    "Place of Death: Place ID\n[GEOCODE]",
    "Place of Death: Formatted Address\n[GEOCODE]",
    "Place of Death: Latitude & Longitude\n[GEOCODE]",
    "Death Date: Month\n[TRANSCRIBED]",
    "Death Date: Day\n[TRANSCRIBED]",
    "Death Date: Year\n[TRANSCRIBED]",
    "Death Date\n[DISPLAYED]",
    "Death Date: ISO\n[SEARCH]",
    "Death Date: Ult Month\n[ADMIN]",
    "Cause of Death\n[TRANSCRIBED]",
    "Cause of Death\n[DISPLAYED]",
    "Undertaker\n[TRANSCRIBED]",
    "Undertaker\n[DISPLAYED]",
    "Remarks\n[TRANSCRIBED]",
    "Remarks\n[DISPLAYED]",
    "Burial Origin\n[ADMIN]",
    "Has Diagram\n[ADMIN]",
    "Transcriber Review\n[ADMIN]",
    "Needs Review\n[ADMIN]",
    "Needs Review Comments\n[ADMIN]"
])
for cell in ws["1:1"]:
    cell.style = header

print('[')
for row in sheet.iter_rows(min_row=args.row_start, values_only=True):

    # if row[INTERMENT_ID] is not None and count < max:
    if count < max:
        # lookup death and interment years if no columns are available
        lookup_years = False
        if args.lookup_years == 'Y':
            lookup_years = True
        i = Interment()
        i.set_previous(previous)
        i.set_id(row[INTERMENT_ID])
        i.set_registry_image_filename_raw(str(row[IMAGE_FILENAME]).strip(), lookup_years)
        # set death day next to last since "removal" status should be determined first

        # expand any year abbreviations
        interment_year_transcribed = row[INTERMENT_YEAR]
        if interment_year_transcribed is not None and (isinstance(interment_year_transcribed, float) or str(interment_year_transcribed).isdigit()):
            if len(str(int(interment_year_transcribed))) == 2:
                interment_year_transcribed = int(year_prefix + str(int(interment_year_transcribed)))

        if lookup_years:
            i.set_interment_date(row[INTERMENT_MONTH], row[INTERMENT_DAY], None)
        else:
            i.set_interment_date(row[INTERMENT_MONTH], row[INTERMENT_DAY], interment_year_transcribed)

        i.set_name_raw(row[NAME])
        i.set_burial_location_lot_raw(row[LOT_LOCATION])
        i.set_burial_location_grave_raw(row[GRAVE_LOCATION])
        i.set_birth_place_raw(row[BIRTH_PLACE])
        i.set_age_years_raw(row[AGE_YEARS])
        i.set_age_months_raw(row[AGE_MONTHS])
        i.set_age_days_raw(row[AGE_DAYS])

        # 2-COLUMN MARITAL STATUS
        if args.marital_status_cols == 2:
            i.set_marital_status_married_raw(row[MARITAL_STATUS_MARRIED])
            i.set_marital_status_single_raw(row[MARITAL_STATUS_SINGLE])
            # merge marital status columns for display
            if i.get_marital_status_married() == 'Not recorded' and i.get_marital_status_single() != 'Not recorded':
                i.set_marital_status(i.get_marital_status_single())
            else:
                i.set_marital_status(i.get_marital_status_married())

        # 1-COLUMN MARITAL STATUS
        if args.marital_status_cols == 1:
            i.set_marital_status_combined_raw(row[MARITAL_STATUS_SINGLE_COLUMN])

        i.set_residence_place_street_raw(row[RESIDENCE_STREET])
        i.set_residence_place_city_raw(row[RESIDENCE_CITY])
        i.set_residence_place_geocode()

        i.set_death_place_raw(row[DEATH_LOCATION])

        i.set_cause_of_death_raw(str(row[CAUSE_OF_DEATH]))
        i.set_undertaker_raw(row[UNDERTAKER])
        i.set_remarks_raw(row[NOTES])

        i.set_has_diagram(row[HAS_DIAGRAM])
        # check if grave location has "see r column"
        if bool(row[GRAVE_LOCATION] and _see_r_column_regex.search(row[GRAVE_LOCATION])):
            i.set_has_diagram('Diagram')

        i.set_transcriber_requests_review(row[NEEDS_REVIEW])

        # expand any year abbreviations
        death_year_transcribed = row[DEATH_YEAR]
        if death_year_transcribed is not None and (isinstance(death_year_transcribed, float) or str(death_year_transcribed).isdigit()):
            if len(str(int(death_year_transcribed))) == 2:
                death_year_transcribed = int(year_prefix + str(int(death_year_transcribed)))

        # set death day next to last since "removal" status should be determined first
        if lookup_years:
            i.set_death_date(row[DEATH_MONTH], row[DEATH_DAY], None)
        else:
            i.set_death_date(row[DEATH_MONTH], row[DEATH_DAY], death_year_transcribed)

        i.set_needs_review_comments()

        # print(i.to_json() + ",")

        # if count == 0:
        #     print(i.to_csv(True))
        # else:
        #     print(i.to_csv(False))

        # EXCEL OUTPUT
        ws.append([
            i.get_id(),
            i.get_registry_image_filename(),
            i.get_interment_date_month_raw(),
            i.get_interment_date_day_raw(),
            i.get_interment_date_year(),
            i.get_interment_date_display(),
            i.get_interment_date_iso(),
            i.get_name_raw(),
            i.get_name_full(),
            i.get_name_last(),
            i.get_name_first(),
            i.get_name_middle(),
            i.get_name_salutation(),
            i.get_name_suffix(),
            i.is_plot_owner(),
            i.get_name_gender_guess(),
            i.get_burial_location_lot_raw(),
            i.get_burial_location_lot(),
            i.get_burial_location_lot_strike(),
            i.get_burial_location_grave_raw(),
            i.get_burial_location_grave(),
            i.get_burial_location_grave_strike(),
            i.get_birth_place_raw(),
            i.get_birth_place_display(),
            i.get_birth_place_full(),
            "FALSE",
            i.get_birth_geo_street_number(),
            i.get_birth_geo_street_name_long(),
            i.get_birth_geo_street_name_short(),
            i.get_birth_geo_neighborhood(),
            i.get_birth_geo_city(),
            i.get_birth_geo_county(),
            i.get_birth_geo_state_short(),
            i.get_birth_geo_state_long(),
            i.get_birth_geo_country_long(),
            i.get_birth_geo_country_short(),
            i.get_birth_geo_zip(),
            i.get_birth_geo_place_id(),
            i.get_birth_geo_formatted_address(),
            str(i.get_birth_geo_location()),
            i.get_age_years_raw(),
            i.get_age_months_raw(),
            i.get_age_days_raw(),
            i.get_age_hours_raw(),
            i.get_age_display(),
            i.get_age_years(),
            i.get_age_months(),
            i.get_age_days(),
            i.get_age_hours(),
            i.get_marital_status_married_raw(),
            i.get_marital_status_single_raw(),
            i.get_marital_status(),
            i.get_residence_place_city_raw(),
            i.get_residence_place_city_full(),
            i.get_residence_place_street_raw(),
            i.get_residence_place_street_full(),
            i.get_residence_place_full(),
            "FALSE",
            i.get_residence_geo_street_number(),
            i.get_residence_geo_street_name_long(),
            i.get_residence_geo_street_name_short(),
            i.get_residence_geo_neighborhood(),
            i.get_residence_geo_city(),
            i.get_residence_geo_county(),
            i.get_residence_geo_state_short(),
            i.get_residence_geo_state_long(),
            i.get_residence_geo_country_long(),
            i.get_residence_geo_country_short(),
            i.get_residence_geo_zip(),
            i.get_residence_geo_place_id(),
            i.get_residence_geo_formatted_address(),
            str(i.get_residence_geo_location()),
            i.get_death_place_raw(),
            i.get_death_place_display(),
            i.get_death_place_full(),
            "FALSE",
            i.get_death_geo_street_number(),
            i.get_death_geo_street_name_long(),
            i.get_death_geo_street_name_short(),
            i.get_death_geo_neighborhood(),
            i.get_death_geo_city(),
            i.get_death_geo_county(),
            i.get_death_geo_state_short(),
            i.get_death_geo_state_long(),
            i.get_death_geo_country_long(),
            i.get_death_geo_country_short(),
            i.get_death_geo_zip(),
            i.get_death_geo_place_id(),
            i.get_death_geo_formatted_address(),
            str(i.get_death_geo_location()),
            i.get_death_date_month_raw(),
            i.get_death_date_day_raw(),
            i.get_death_date_year(),
            i.get_death_date_display(),
            i.get_death_date_iso(),
            i.get_ultimate_month(),
            i.get_cause_of_death_raw(),
            i.get_cause_of_death_display(),
            i.get_undertaker_raw(),
            i.get_undertaker_display(),
            i.get_remarks_raw(),
            i.get_remarks_display(),
            i.get_burial_origin(),
            i.get_has_diagram(),
            i.get_transcriber_requests_review(),
            i.get_needs_review(),
            i.get_needs_review_comments()
        ])

        # add image links
        ws.cell(row=count, column=2).hyperlink = i.get_registry_image_link()
        ws.cell(row=count, column=2).value = i.get_registry_image_filename()
        ws.cell(row=count, column=2).style = "Hyperlink"

        # add comments
        # idcell = ws.cell(row=count, column=1)
        # idcell.comment = Comment(u'This is the comment', u'Comment Author', )

        # needs review fill color
        fill = PatternFill("solid", fgColor="FFD1DC")
        for cell in ws[count:count]:
            cell.style = content
            if i.get_needs_review():
                cell.fill = fill

        previous = i
        count += 1
print(']')

# hide all geocode fields
for col in ['Y', 'BE', 'BW', 'BX',
            'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN',
            'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT',
            'BY', 'BZ', 'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL']:
    ws.column_dimensions[col].hidden = True

# hide iso dates
# for col in ['G', 'CQ']:
#     ws.column_dimensions[col].hidden = True

# hide results of name parsing
for col in ['J', 'K', 'L', 'M', 'N']:
    ws.column_dimensions[col].hidden = True

# hide gender guess
for col in ['P']:
    ws.column_dimensions[col].hidden = True

volume_number_str = str(args.vol) if args.vol >= 9 else '0' + str(args.vol)
wb.save("excel/output/Volume_" + volume_number_str + "_processed.xlsx")