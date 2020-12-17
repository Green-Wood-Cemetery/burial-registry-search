import json
import re
import dateparser
import requests
import logging
from nameparser import HumanName
from openpyxl import load_workbook
import argparse
import sqlite3
from inflection import parameterize
from os import path
import time
import googlemaps

connection = sqlite3.connect("gender/gender.db")
cursor = connection.cursor()

# command line arguments
parser = argparse.ArgumentParser(description='Spreadsheet to JSON')
parser.add_argument('-input', type=str, help='spreadsheet in xslx format')
parser.add_argument('-key', type=str, help='google api key')
parser.add_argument('-sheet', type=str, help='worksheet to transform')
parser.add_argument('-vol', type=str, help='registry volume number')
args = parser.parse_args()
workbook = load_workbook(filename=args.input, data_only=True)

GOOGLE_API_KEY = args.key
gmaps = googlemaps.Client(key=GOOGLE_API_KEY)

# sheet = workbook.active
sheet = workbook[args.sheet]

# logging config
timestr = time.strftime("%Y%m%d-%H%M%S")
logging.basicConfig(filename='logs/import-volume-' + args.vol + '-' + timestr + '.csv', filemode='a', format='%(message)s')
logging.warning('"volume","interment id","message"')

do_geocode_birth = True
do_geocode_residence = True
do_geocode_death = True
cemetery = "Green-Wood Cemetery, Brooklyn, NY, USA"


# load synonym dictionaries
with open('dictionaries/cause-of-death.json') as f:
    cause_of_death_dict = json.load(f)
with open('dictionaries/marital-status.json') as g:
    marital_status_dict = json.load(g)
with open('dictionaries/city-to-state.json') as g:
    city_state_dict = json.load(g)
with open('dictionaries/states.json') as g:
    state_dict = json.load(g)
with open('dictionaries/name-abbrev.json') as g:
    name_abbrev_dict = json.load(g)

# spreadsheet column constants
IMAGE_FILENAME = 0
# col 1 contains image links for QA purposes only
INTERMENT_ID = 2

INTERMENT_MONTH = 3
INTERMENT_DAY = 4
INTERMENT_YEAR = 5

NAME_SALUTATION = 6
NAME_FIRST = 7
NAME_MIDDLE = 8
NAME_LAST = 9
NAME_INFANT = 10

BURIAL_LOCATION_CURRENT_LOT = 11
BURIAL_LOCATION_CURRENT_GRAVE = 12
BURIAL_LOCATION_ORIGINAL_LOT = 13
BURIAL_LOCATION_ORIGINAL_GRAVE = 14

BIRTH_CITY = 15
BIRTH_STATE = 16
BIRTH_COUNTRY = 17

AGE_YEARS = 18
AGE_MONTHS = 19
AGE_DAYS = 20

MARITAL_STATUS = 21

RESIDENCE_STREET = 22
RESIDENCE_CITY = 23
RESIDENCE_STATE = 24
RESIDENCE_COUNTRY = 25

DEATH_LOCATION = 26
DEATH_STREET = 27
DEATH_CITY = 28
DEATH_STATE = 29
DEATH_COUNTRY = 30

DEATH_MONTH = 31
DEATH_DAY = 32
DEATH_YEAR = 33

CAUSE_OF_DEATH = 34

REMOVAL_FROM = 35

UNDERTAKER = 36

REMOVED = 37
REMOVED_DATE = 38

NOTES = 39

TAGS = 40


# =====================================================================================================================
# GOOGLE GEOCODE API LOOKUP
# =====================================================================================================================
def get_google_geocode_results(address_or_zipcode):
    base_url = "https://maps.googleapis.com/maps/api/geocode/json"
    endpoint = f"{base_url}?address={address_or_zipcode}&key={api_key}"
    # see how our endpoint includes our API key? Yes this is yet another reason to restrict the key
    r = requests.get(endpoint)
    if r.status_code not in range(200, 299):
        return None, None
    try:
        '''
        This try block in case any of our inputs are invalid. This is done instead
        of actually writing out handlers for all kinds of responses.
        '''
        results = r.json()['results'][0]
    except:
        pass
    return results


def get_geocode_dict():
    d = {}
    keys = [
        'place_geocode_input',
        'geo_lat', 'geo_lng',
        'geo_formatted_address',
        'geo_street_number',
        'geo_street_name_long',
        'geo_street_name_short',
        'geo_neighborhood',
        'geo_country_long',
        'geo_country_short',
        'geo_city',
        'geo_state_long',
        'geo_state_short',
        'geo_county',
        'geo_zip',
        'geo_other',
        'google_place_id'
    ]
    # initialize empty dictionary
    for i in keys:
        d[i] = ''
    return d

# =====================================================================================================================
# GEOCODE PLACE
# =====================================================================================================================
def geocode_place(vol, id, type, place):

    d = get_geocode_dict()
    # return empty dictionary if no place is specified
    if place == '':
        return d
    else:
        d['place_geocode_input'] = place

    geocode_filename = 'json/places/' + parameterize(place.lower().strip(), separator="_") + '.json'
    if place != '' and path.exists(geocode_filename) is False:

        # geocode_results = get_google_geocode_results(place)
        gmaps_result = gmaps.geocode(place)
        if len(gmaps_result) != 0:
            geocode_results = gmaps_result[0]
        else:
            geocode_results = None

        if geocode_results is not None:
            d['google_place_id'] = geocode_results['place_id']
            d['place_geocode_input'] = place
            d['geo_lat'] = geocode_results['geometry']['location']['lat']
            d['geo_lng'] = geocode_results['geometry']['location']['lng']
            d['geo_formatted_address'] = geocode_results['formatted_address']
            for component in geocode_results['address_components']:
                for geo_type in component['types']:
                    if geo_type == 'street_number':
                        d['geo_street_number'] = component['short_name']
                    if geo_type == 'route':
                        d['geo_street_name_long'] = component['long_name']
                        d['geo_street_name_short'] = component['short_name']
                    if geo_type == 'neighborhood':
                        d['geo_neighborhood'] = component['long_name']
                    if geo_type == 'country':
                        d['geo_country_short'] = component['short_name']
                        d['geo_country_long'] = component['long_name']
                    if geo_type == 'sublocality':
                        d['geo_city'] = component['long_name']
                    if geo_type == 'locality':
                        d['geo_city'] = component['long_name']
                    if geo_type == 'administrative_area_level_1':
                        d['geo_state_long'] = component['long_name']
                        d['geo_state_short'] = component['short_name']
                    if geo_type == 'administrative_area_level_2':
                        d['geo_county'] = component['long_name']
                    if geo_type == 'postal_code':
                        d['geo_zip'] = component['long_name']
                    if geo_type == 'establishment':
                        d['geo_other'] = component['long_name']
                    if geo_type == 'natural_feature':
                        d['geo_other'] = component['long_name']
            try:
                f = open(geocode_filename, 'w')
                json.dump(d, f, indent=4, sort_keys=True)
                f.close()
                return d
            except Exception as e:
                logging.fatal(getattr(e, 'message', repr(e)))
        else:
            # logging.warning("VOLUME " + vol + " INTERMENT ID " + str(int(id)) + " unable geocode " + type + " place: " + place)
            logging.warning('"' + vol + '","' + str(int(id)) + '","' + "Unable geocode " + type + " place: " + place + '"')
            # serialize empty geo place
            f = open(geocode_filename, 'w')
            json.dump(d, f, indent=4, sort_keys=True)
            f.close()
            return d
    else:
        # return serialized geocodes
        if path.exists(geocode_filename):
            with open(geocode_filename) as json_file:
                place_json = json.load(json_file)
                if place_json["google_place_id"] == "":
                    # logging.warning("VOLUME " + vol + " INTERMENT ID " + str(int(id)) + " unable to geocode " + type + " place: " + place)
                    logging.warning('"' + vol + '","' + str(int(id)) + '","' + "Unable geocode " + type + " place: " + place + '"')
                return place_json
        else:
            return d

# =====================================================================================================================
# IS FLOAT?
# =====================================================================================================================
def isfloat(value):
  try:
    float(value)
    return True
  except ValueError:
    return False

# =====================================================================================================================
# MAIN
# =====================================================================================================================
interments = []

# Using the values_only because you want to return the cells' values
for row in sheet.iter_rows(min_row=3, values_only=True):

    if row[INTERMENT_ID] is not None:

        # --- INTERMENT ID
        interment_id = row[INTERMENT_ID]

        # --- TAGS
        tags = []
        if len(row) > TAGS:
            if row[TAGS] is not None:
                tags = re.findall(r'"(.*?)"', row[TAGS])

        # --- REGISTRY IMAGE FILENAME
        image_filename = str(row[IMAGE_FILENAME]).strip()

        # --- PARSE REGISTRY VOL AND PAGE
        try:
            m = re.search('[Vv]olume\s+(\d+)_(\d+)', image_filename)
            if m == None:
                # logging.warning("VOLUME " + str(registry_volume) + " INTERMENT ID " + str(int(interment_id)) + " unable to parse image filename: " + image_filename)
                logging.warning('"' + str(registry_volume) + '","' + str(int(interment_id)) + '","' + "Unable to parse image filename: " + image_filename + '"')
                exit()
            registry_volume = m.group(1)
            registry_page = m.group(2)
            image_filename = "Volume " + registry_volume + "_" + registry_page
            # todo; check if image exists on server
        except re.error:
            # logging.warning("VOLUME " + str(registry_volume) + " INTERMENT ID " + str(int(interment_id)) + " unable to parse image filename: " + image_filename)
            logging.warning('"' + str(registry_volume) + '","' + str(int(interment_id)) + '","' + "Unable to parse image filename: " + image_filename + '"')
            exit()

        # --- INTERMENT DATE
        interment_date_display = ''
        interment_date_iso = ''
        interment_year = ''
        # interment month
        if row[INTERMENT_MONTH] is not None and row[INTERMENT_MONTH] != '':
            interment_date_display += row[INTERMENT_MONTH] + " "
        # interment day
        if row[INTERMENT_DAY] is not None and row[INTERMENT_DAY] != '':
            interment_date_display += str(int(row[INTERMENT_DAY])) + ", "
        # interment year
        if row[INTERMENT_YEAR] is not None and row[INTERMENT_YEAR] != '':
            interment_date_display += str(int(row[INTERMENT_YEAR]))
            interment_year = int(row[INTERMENT_YEAR])
            if interment_year > 2020:
                # logging.warning("VOLUME " + str(registry_volume) + " INTERMENT ID " + str(int(interment_id)) + " has an interment_year greater than 2020: " + str(interment_year) )
                logging.warning('"' + str(registry_volume) + '","' + str(int(interment_id)) + '","' + "Has an interment_year greater than 2020: " + str(interment_year) + '"')
        interment_date = ''
        if interment_date_display != '':
            dt = dateparser.parse(interment_date_display)
            if dt is not None:
                interment_date_iso = dt.strftime("%Y-%m-%d")
            else:
                # logging.warning("VOLUME " + str(registry_volume) + " INTERMENT ID " + str(int(interment_id)) + " unable to parse interment date: " + interment_date_display )
                logging.warning('"' + str(registry_volume) + '","' + str(int(interment_id)) + '","' + "Unable to parse interment date: " + interment_date_display + '"')

        # --- NAME
        name_salutation = ''
        name_first = ''
        name_first_gender = 'Unknown'
        name_middle = ''
        name_last = ''
        name_infant = ''
        name_full = ''

        if row[NAME_SALUTATION] is not None:
            name_salutation = str(row[NAME_SALUTATION]).strip()

        if row[NAME_FIRST] is not None:
            name_first = str(row[NAME_FIRST]).strip()

            # try to guess gender using first namegenderpro database
            # just grab first part of any first name (eg: Mary Jane)
            name_first_gender_temp = name_first.split(' ')[0]
            gender_row = cursor.execute("SELECT gender from namegenderpro where name = '" + name_first_gender_temp + "' COLLATE NOCASE").fetchone()
            if gender_row:
                    name_first_gender = gender_row[0]

        if row[NAME_MIDDLE] is not None:
            name_middle = str(row[NAME_MIDDLE]).strip()

        if row[NAME_LAST] is not None:
            name_last = str(row[NAME_LAST]).strip()

        name = HumanName(name_salutation + " " + name_first + " " + name_middle + " " + name_last)
        name_full = name.full_name

        if row[NAME_INFANT] is not None:
          name_infant = row[NAME_INFANT]
          if name_infant != '':
              name_full += " (" + name_infant + ")"
              if row[NAME_LAST] is None:
                logging.warning('"' + str(registry_volume) + '","' + str(int(interment_id)) + '","' + "Has an infant name but is missing last name" + '"')

        #--- expand abbreviated first names in tags using name-abbrev dictionary
        for key in name_abbrev_dict.keys():
            if key == name_first.lower():
                if name_abbrev_dict[key] not in tags:
                    tags.append(name_abbrev_dict[key])

        #--- BURIAL LOCATION
        burial_location_current_lot = ''
        burial_location_current_grave = ''
        burial_location_previous_lot = ''
        burial_location_previous_grave = ''
        if row[BURIAL_LOCATION_CURRENT_LOT] is not None:
            burial_location_current_lot = row[BURIAL_LOCATION_CURRENT_LOT]
        if row[BURIAL_LOCATION_CURRENT_GRAVE] is not None:
            burial_location_current_grave = row[BURIAL_LOCATION_CURRENT_GRAVE]
        if row[BURIAL_LOCATION_ORIGINAL_LOT] is not None:
            burial_location_previous_lot = row[BURIAL_LOCATION_ORIGINAL_LOT]
        if row[BURIAL_LOCATION_ORIGINAL_GRAVE] is not None:
            burial_location_previous_grave = row[BURIAL_LOCATION_ORIGINAL_GRAVE]

        # --- PLACE OF BIRTH
        birth_city = ''
        birth_state = ''
        birth_country = ''
        birth_place_full = ''
        birth_geo_street_number = ''
        birth_geo_street_name_long = ''
        birth_geo_street_name_short = ''
        birth_geo_neighborhood = ''
        birth_geo_country_short = ''
        birth_geo_country_long = ''
        birth_geo_city = ''
        birth_geo_state_long = ''
        birth_geo_state_short = ''
        birth_geo_county = ''
        birth_geo_zip = ''
        birth_geo_lat = None
        birth_geo_lng = None
        birth_geo_formatted_address = ''

        # BIRTH CITY
        if row[BIRTH_CITY] is not None:
            birth_city = str(row[BIRTH_CITY]).strip()
            # city is really a state, and state is empty, move it over
            for key in state_dict.keys():
                if key == birth_city.lower() and row[BIRTH_STATE] is None:
                    birth_city = ''
                    birth_state = key
            # long s variant of massachusetts
            if 'mafs' == birth_city.lower() and row[BIRTH_STATE] is None:
                birth_city = ''
                birth_state = 'Massachusetts'
                if row[BIRTH_COUNTRY] is None:
                    birth_country = 'United States'

        # BIRTH STATE
        if row[BIRTH_STATE] is not None:
            birth_state = str(row[BIRTH_STATE]).strip()

        # BIRTH COUNTRY
        if row[BIRTH_COUNTRY] is not None:
            birth_country = str(row[BIRTH_COUNTRY])

        # BIRTH PLACE FULL - concatenate all of the available fields
        birth_place_full = (birth_city + " " + birth_state + " " + birth_country).strip()
        ' '.join(birth_place_full.split())

        # GEOCODE: BIRTH PLACE
        if do_geocode_birth:
            birth_place_geocoded = geocode_place(registry_volume, interment_id, 'birth', birth_place_full)
        else:
            birth_place_geocoded = get_geocode_dict()

        # --- AGE
        age_years = None
        age_months = None
        age_days = None
        age_full = ''
        if row[AGE_YEARS] is not None and row[AGE_YEARS] != '':
            age_years = int(row[AGE_YEARS])
            age_full += str(int(age_years)) + " years"
            if age_years > 120:
                # logging.warning("VOLUME " + str(registry_volume) + " INTERMENT ID " + str(int(interment_id)) + " has an age greater than 120 years: " + str(int(age_years)))
                logging.warning('"' + str(registry_volume) + '","' + str(int(interment_id)) + '","' + "Has an age greater than 120 years: " + str(int(age_years)) + '"')

        if row[AGE_MONTHS] is not None and row[AGE_MONTHS] != '':
            age_months = row[AGE_MONTHS]
            if age_full != '':
                age_full += ", "
            age_full += str(int(age_months)) + " months"
        if row[AGE_DAYS] is not None and row[AGE_DAYS] != '':
            age_days = str(row[AGE_DAYS])
            age_days = re.sub('\s*1/2', '.5', age_days)
            age_days = re.sub('\s*3/4', '.75', age_days)
            age_days = re.sub('\d+\s+Hrs?', '0', age_days)
            age_days = re.sub('\w+\s+[Hh]ours?', '0', age_days)
            if isfloat(age_days):
                age_days = float(age_days)
                if age_full != '':
                    age_full += ", "
                age_full += str(age_days) + " days"

        # --- MARITAL STATUS
        valid_status = ['Not recorded', 'Married', 'Single', 'Widow', 'Unknown']
        marital_status = 'Not recorded'
        if row[MARITAL_STATUS] is not None and row[MARITAL_STATUS] != '':
            marital_status = str(row[MARITAL_STATUS]).strip().capitalize()
        # normalize variants
        for key in marital_status_dict.keys():
            if key == marital_status.lower():
                marital_status = (marital_status_dict[key])
        if marital_status not in valid_status:
            # logging.warning("VOLUME " + str(registry_volume) + " INTERMENT ID " + str(int(interment_id)) + " has a marital status of: " + marital_status)
            logging.warning('"' + str(registry_volume) + '","' + str(int(interment_id)) + '","' + "Has a marital status of: " + marital_status + '"')

        # --- PLACE OF RESIDENCE
        residence_street = ''
        residence_city = ''
        residence_state = ''
        residence_country = ''
        residence_place_full = ''
        residence_geo_street_number = ''
        residence_geo_street_name_long = ''
        residence_geo_street_name_short = ''
        residence_geo_neighborhood = ''
        residence_geo_country_short = ''
        residence_geo_country_long = ''
        residence_geo_city = ''
        residence_geo_state_long = ''
        residence_geo_state_short = ''
        residence_geo_county = ''
        residence_geo_zip = ''
        residence_geo_lat = None
        residence_geo_lng = None
        residence_geo_formatted_address = ''
        if row[RESIDENCE_STREET] is not None:
            residence_street = str(row[RESIDENCE_STREET]).strip()
        if row[RESIDENCE_CITY] is not None:
            residence_city = str(row[RESIDENCE_CITY]).strip()

            # long s variant of massachusetts
            if 'mafs' == residence_city.lower() and row[RESIDENCE_STATE] is None:
                residence_city = ''
                residence_state = 'Massachusetts'
                if row[RESIDENCE_COUNTRY] is None:
                    residence_country = 'United States'

            # infer state in some cases using a mapping dictionary
            if row[RESIDENCE_STATE] is None:
                for key in city_state_dict.keys():
                    if key == residence_city.lower():
                        residence_state = (city_state_dict[key])
        if row[RESIDENCE_STATE] is not None:
            residence_state = str(row[RESIDENCE_STATE]).strip()
        if row[RESIDENCE_COUNTRY] is not None:
            residence_country = str(row[RESIDENCE_COUNTRY]).strip()

        residence_place_full = (residence_street + " " + residence_city + " " + residence_state + " " + residence_country).strip()
        ' '.join(residence_place_full.split())

        # GEOCODE: RESIDENCE PLACE
        if do_geocode_residence:
            residence_place_geocoded = geocode_place(registry_volume, interment_id, 'residence', residence_place_full)
        else:
            residence_place_geocoded = get_geocode_dict()

        # --- PLACE OF DEATH
        death_location = ''
        death_street = ''
        death_city = ''
        death_state = ''
        death_country = ''
        death_place_full = ''
        death_geo_street_number = ''
        death_geo_street_name_long = ''
        death_geo_street_name_short = ''
        death_geo_neighborhood = ''
        death_geo_country_short = ''
        death_geo_country_long = ''
        death_geo_city = ''
        death_geo_state_long = ''
        death_geo_state_short = ''
        death_geo_county = ''
        death_geo_zip = ''
        death_geo_lat = None
        death_geo_lng = None
        death_geo_formatted_address = ''
        if row[DEATH_LOCATION] is not None:
            death_location = str(row[DEATH_LOCATION]).strip()
        if row[DEATH_STREET] is not None:
            death_street = str(row[DEATH_STREET]).strip()
        if row[DEATH_CITY] is not None:
            death_city = str(row[DEATH_CITY]).strip()

            # long s variant of massachusetts
            if 'mafs' == death_city.lower() and row[DEATH_STATE] is None:
                death_city = ''
                death_state = 'Massachusetts'
                if row[DEATH_COUNTRY] is None:
                    death_country = 'United States'

            # infer state in some cases using a mapping dictionary
            if row[DEATH_STATE] is None:
                for key in city_state_dict.keys():
                    if key == death_city.lower():
                        death_state = (city_state_dict[key])
        if row[DEATH_STATE] is not None:
            death_state = str(row[DEATH_STATE]).strip()
        if row[DEATH_COUNTRY] is not None:
            death_country = str(row[DEATH_COUNTRY]).strip()
        death_place_full = (death_location + " " + death_street + " " + death_city + " " + death_state + " " + death_country).strip()
        ' '.join(death_place_full.split())

        # GEOCODE: DEATH PLACE
        if do_geocode_death:
            death_place_geocoded = geocode_place(registry_volume, interment_id, 'death', death_place_full)
        else:
            death_place_geocoded = get_geocode_dict()

        # --- DEATH DATE
        death_date_display = ''
        death_date_iso = ''
        death_year = ''
        # death month
        if row[DEATH_MONTH] is not None and row[DEATH_MONTH] != '':
            death_date_display += row[DEATH_MONTH] + " "
        # death day
        if row[DEATH_DAY] is not None and row[DEATH_DAY] != '':
            # correct capital letter O used instead of zero
            death_day = str(row[DEATH_DAY]).strip()
            death_day = death_day.replace("O","0").replace(".0", "")
            death_date_display += death_day + ", "
        # death year
        if row[DEATH_YEAR] is not None and row[DEATH_YEAR] != '':
            death_date_display += str(int(row[DEATH_YEAR]))
            death_year = int(row[DEATH_YEAR])
            if death_year > 2020:
                # logging.warning("VOLUME " + str(registry_volume) + " INTERMENT ID " + str(int(interment_id)) + " has a death year greater than 2020: " + str(death_year) )
                logging.warning('"' + str(registry_volume) + '","' + str(int(interment_id)) + '","' + "Has a death year greater than 2020: " + str(death_year) + '"')
        death_date = ''
        if death_date_display != '':
            dt = dateparser.parse(death_date_display)
            if dt is not None:
                death_date_iso = dt.strftime("%Y-%m-%d")
            else:
                # logging.warning("VOLUME " + str(registry_volume) + " INTERMENT ID " + str(int(interment_id)) + " unable to parse death date: " + death_date_display )
                logging.warning('"' + str(registry_volume) + '","' + str(int(interment_id)) + '","' + "Unable to parse death date: " + death_date_display + '"')
        # --- CAUSE OF DEATH
        cause_of_death = ''
        if row[CAUSE_OF_DEATH] is not None:
            cause_of_death = row[CAUSE_OF_DEATH].strip()

        # add synonyms as tags
        for key in cause_of_death_dict.keys():
            if key == cause_of_death.lower():
                tags.append(cause_of_death_dict[key])

        # --- REMOVAL FROM
        removal_from = ''
        if row[REMOVAL_FROM] is not None:
            removal_from = str(row[REMOVAL_FROM]).strip()

        # --- UNDERTAKER
        undertaker = ''
        if row[UNDERTAKER] is not None:
            undertaker = str(row[UNDERTAKER]).strip()

        # --- REMOVED
        removed = False
        if row[REMOVED] is not None:
            removed = True

        # --- REMOVED DATE
        removed_date = ''
        if row[REMOVED_DATE] is not None:
            removed_date = str(row[REMOVED_DATE]).strip()

        # --- NOTES
        notes = ''
        if row[NOTES] is not None:
            notes = str(row[NOTES]).strip()

        interment = {
            "cemetery": cemetery,
            "interment_id": interment_id,
            "image_filename": image_filename,
            "registry_volume": registry_volume,
            "interment_date_display" : interment_date_display,
            "interment_date_iso": interment_date_iso,
            "interment_year": interment_year,
            "name_salutation": name_salutation,
            "name_first": name_first,
            "name_middle": name_middle,
            "name_last": name_last,
            "name_full": name_full,
            "name_infant": name_infant,
            "gender_guess": name_first_gender,
            "burial_location_current_lot": burial_location_current_lot,
            "burial_location_current_grave": burial_location_current_grave,
            "burial_location_previous_lot": burial_location_previous_lot,
            "burial_location_previous_grave": burial_location_previous_grave,
            "birth_city": birth_city,
            "birth_state": birth_state,
            "birth_country": birth_country,
            "birth_place_full": birth_place_full,
            "birth_geo_location": {"lat": birth_place_geocoded['geo_lat'],"lon": birth_place_geocoded['geo_lng']},
            "birth_geo_street_number": birth_place_geocoded['geo_street_number'],
            "birth_geo_street_name_long": birth_place_geocoded['geo_street_name_long'],
            "birth_geo_street_name_short": birth_place_geocoded['geo_street_name_short'],
            "birth_geo_neighborhood": birth_place_geocoded['geo_neighborhood'],
            "birth_geo_city": birth_place_geocoded['geo_city'],
            "birth_geo_county": birth_place_geocoded['geo_county'],
            "birth_geo_state_short": birth_place_geocoded['geo_state_short'],
            "birth_geo_state_long": birth_place_geocoded['geo_state_long'],
            "birth_geo_country_long": birth_place_geocoded['geo_country_long'],
            "birth_geo_country_short": birth_place_geocoded['geo_country_short'],
            "birth_geo_zip": birth_place_geocoded['geo_zip'],
            "birth_geo_place_id": birth_place_geocoded['google_place_id'],
            "birth_geo_formatted_address": birth_place_geocoded['geo_formatted_address'],
            "age_years": age_years,
            "age_months": age_months,
            "age_days": age_days,
            "age_full": age_full,
            "marital_status": marital_status,
            "residence_street": residence_street,
            "residence_city": residence_city,
            "residence_state": residence_state,
            "residence_country": residence_country,
            "residence_place_full": residence_place_full,
            "residence_geo_location": {"lat": residence_place_geocoded['geo_lat'],"lon": residence_place_geocoded['geo_lng']},
            "residence_geo_street_number": residence_place_geocoded['geo_street_number'],
            "residence_geo_street_name_long": residence_place_geocoded['geo_street_name_long'],
            "residence_geo_street_name_short": residence_place_geocoded['geo_street_name_short'],
            "residence_geo_neighborhood": residence_place_geocoded['geo_neighborhood'],
            "residence_geo_city": residence_place_geocoded['geo_city'],
            "residence_geo_county": residence_place_geocoded['geo_county'],
            "residence_geo_state_short": residence_place_geocoded['geo_state_short'],
            "residence_geo_state_long": residence_place_geocoded['geo_state_long'],
            "residence_geo_country_long": residence_place_geocoded['geo_country_long'],
            "residence_geo_country_short": residence_place_geocoded['geo_country_short'],
            "residence_geo_zip": residence_place_geocoded['geo_zip'],
            "residence_geo_place_id": residence_place_geocoded['google_place_id'],
            "residence_geo_formatted_address": residence_place_geocoded['geo_formatted_address'],
            "death_location": death_location,
            "death_street": death_street,
            "death_city": death_city,
            "death_state": death_state,
            "death_country": death_country,
            "death_place_full": death_place_full,
            "death_geo_location": {"lat": death_place_geocoded['geo_lat'],"lon": death_place_geocoded['geo_lng']},
            "death_geo_street_number": death_place_geocoded['geo_street_number'],
            "death_geo_street_name_long": death_place_geocoded['geo_street_name_long'],
            "death_geo_street_name_short": death_place_geocoded['geo_street_name_short'],
            "death_geo_neighborhood": death_place_geocoded['geo_neighborhood'],
            "death_geo_city": death_place_geocoded['geo_city'],
            "death_geo_county": death_place_geocoded['geo_county'],
            "death_geo_state_short": death_place_geocoded['geo_state_short'],
            "death_geo_state_long": death_place_geocoded['geo_state_long'],
            "death_geo_country_long": death_place_geocoded['geo_country_long'],
            "death_geo_country_short": death_place_geocoded['geo_country_short'],
            "death_geo_zip": death_place_geocoded['geo_zip'],
            "death_geo_place_id": death_place_geocoded['google_place_id'],
            "death_geo_formatted_address": death_place_geocoded['geo_formatted_address'],
            "death_date_display" : death_date_display,
            "death_date_iso": death_date_iso,
            "death_year": death_year,
            "cause_of_death": cause_of_death,
            "removal_from": removal_from,
            "undertaker": undertaker,
            "removed": removed,
            "removed_date": removed_date,
            "note": notes,
            "tags": tags
        }
        # make sure you can json serialize, otherwise dump error and interment
        try:
            json_temp = json.dumps(interment)
            interments.append(interment)
        except Exception as e:
            logging.fatal(getattr(e, 'message', repr(e)))
            logging.fatal(interment)
            # exit()

print(json.dumps(interments))
connection.close()